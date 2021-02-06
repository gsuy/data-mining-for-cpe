import random
import matplotlib.pyplot as plt
import openpyxl
import math
import numpy as np

File = openpyxl.load_workbook('Data11.xlsx')
sheet = File.active

class NeuralNetwork():
    learningRate = None
    layerSize = [10,10,4]
    w_Layer = [None,None,None]
    bias = [None,None,None]
    y = [None,None,None]
    input_ = []
    output = []
    sizeInput = 13
    def __init__(self,l):
        self.learningRate = l
        add = []
        for i in range(self.sizeInput):
            temp = []
            for i2 in range(self.layerSize[0]):
                temp.append(random.uniform(0.1, 1.0))
            add.append(temp)
        a = np.array(add)
        self.w_Layer[0] = a
        
        add = []
        for i3 in range(self.layerSize[0]):
            temp = []
            for i4 in range(self.layerSize[1]):
                temp.append(random.uniform(0.1, 1.0))
            add.append(temp)
        a = np.array(add)
        self.w_Layer[1] = a

        add = []
        for i5 in range(self.layerSize[1]):
            temp = []
            for i6 in range(self.layerSize[2]):
                temp.append(random.uniform(0.1, 1.0))
            add.append(temp)
        a = np.array(add)
        self.w_Layer[2] = a

        for k in range(len(self.layerSize)):
            temp = []
            for kk in range(self.layerSize[k]):
                temp.append(random.uniform(0.1, 1.0))
            a = np.array(temp)
            self.bias[k] = a
               
    def sigmoid(self,i_):
        temp = []
        for hh in i_:
            ans = 1/(1 + math.exp(-hh))
            temp.append(ans)
        return np.array(temp)
        
    def forward(self,input__):
        self.input_ = np.array(input__)
        self.y[0] = self.sigmoid(self.input_.dot(self.w_Layer[0]) + self.bias[0])
        self.y[1] = self.sigmoid(self.y[0].dot(self.w_Layer[1]) + self.bias[1])
        self.y[2] = self.sigmoid(self.y[1].dot(self.w_Layer[2]) + self.bias[2])
        self.output = self.y[2]
    def backward(self,output,real):
        error = real - output
        gradient_output = []
        deltaW = [None,None,None]
        #หา gradient ของ node output
        for i in range(len(error)):
            gradient_output.append(error[i]*output[i]*(1-output[i]))

        add = []
        for n in range(len(self.w_Layer[2])):
            temp = []
            for nn in gradient_output:
                temp.append(-self.learningRate * nn * self.y[1][n])
            add.append(temp)
        deltaW[2] = add

        #หา gradient ของ node layer2
        gradient_Layer2 = []
        for b in range(len(self.y[1])):
            temp = (self.w_Layer[2][b].dot(gradient_output))*(1-self.y[1][b])*self.y[1][b]
            gradient_Layer2.append(temp)

        add = []
        for j in range(len(self.w_Layer[1])):
            temp = []
            for jj in gradient_Layer2:
                temp.append(-self.learningRate * jj *self.y[0][j])
            add.append(temp)
        deltaW[1] = add
        

        #หา gradient ของ node layer1
        gradient_Layer1 = []
        for s in range(len(self.y[0])):
            temp = (self.w_Layer[1][s].dot(gradient_Layer2))*(1-self.y[0][s])*self.y[0][s]
            gradient_Layer1.append(temp)

        add = []
        for x in range(len(self.w_Layer[0])):
            temp = []
            for xx in gradient_Layer1:
                temp.append(-self.learningRate * xx *self.input_[x])
            add.append(temp)
        deltaW[0] = add

        #ปรับ bias
        deltaB = np.array(gradient_output).dot(-self.learningRate)
        self.bias[2] = self.bias[2] + deltaB
        deltaB = np.array(gradient_Layer2).dot(-self.learningRate)
        self.bias[1] = self.bias[1] + deltaB
        deltaB = np.array(gradient_Layer1).dot(-self.learningRate)
        self.bias[0] = self.bias[0] + deltaB

        #ปรับ weight
        deltaW = np.array(deltaW)
        self.w_Layer = self.w_Layer + deltaW

    def train(self,in_,out_):
        self.forward(in_)
        self.backward(self.output,out_)
    def test(self,in_put):
        self.forward(in_put)
        return self.output 


class NeuralNetwork_():
    learningRate = None
    layerSize = [10,4]
    w_Layer = [None,None]
    bias = [None,None]
    y = [None,None]
    input_ = []
    output = []
    sizeInput = 13
    def __init__(self,l,ll):
        self.learningRate = l
        self.layerSize[0] = ll
        #----------------------------------------------------------
        # สร้าง egde ที่มี w เริ่มต้นเป็น 1
        add = []
        for i in range(self.sizeInput):
            temp = []
            for i2 in range(self.layerSize[0]):
                temp.append(random.uniform(0.1, 1.0))
            add.append(temp)
        a = np.array(add)
        self.w_Layer[0] = a
        
        add = []
        for i3 in range(self.layerSize[0]):
            temp = []
            for i4 in range(self.layerSize[1]):
                temp.append(random.uniform(0.1, 1.0))
            add.append(temp)
        a = np.array(add)
        self.w_Layer[1] = a

        # add = []
        # for i5 in range(self.layerSize[1]):
        #     temp = []
        #     for i6 in range(self.layerSize[2]):
        #         temp.append(random.uniform(0.1, 1.0))
        #     add.append(temp)
        # a = np.array(add)
        # self.w_Layer[2] = a

        #สร้าง bias เริ่มต้นเป็น 1
        for k in range(len(self.layerSize)):
            temp = []
            for kk in range(self.layerSize[k]):
                temp.append(random.uniform(0.1, 1.0))
            a = np.array(temp)
            self.bias[k] = a
               
    def sigmoid(self,i_):
        temp = []
        for hh in i_:
            ans = 1/(1 + math.exp(-hh))
            temp.append(ans)
        return np.array(temp)
        
    def forward(self,input__):
        self.input_ = np.array(input__)
        self.y[0] = self.sigmoid(self.input_.dot(self.w_Layer[0]) + self.bias[0])
        self.y[1] = self.sigmoid(self.y[0].dot(self.w_Layer[1]) + self.bias[1])
        # self.y[2] = self.sigmoid(self.y[1].dot(self.w_Layer[2]) + self.bias[2])
        self.output = self.y[1]
    def backward(self,output,real):
        error = real - output
        gradient_output = []
        deltaW = [None,None]
        #หา gradient ของ node output
        for i in range(len(error)):
            gradient_output.append(error[i]*output[i]*(1-output[i]))

        add = []
        for n in range(len(self.w_Layer[1])):
            temp = []
            for nn in gradient_output:
                temp.append(-self.learningRate * nn * self.y[0][n])
            add.append(temp)
        deltaW[1] = add

        #หา gradient ของ node layer2
        gradient_Layer2 = []
        for b in range(len(self.y[0])):
            temp = (self.w_Layer[1][b].dot(gradient_output))*(1-self.y[0][b])*self.y[0][b]
            gradient_Layer2.append(temp)

        add = []
        for j in range(len(self.w_Layer[0])):
            temp = []
            for jj in gradient_Layer2:
                temp.append(-self.learningRate * jj *self.input_[j])
            add.append(temp)
        deltaW[0] = add

        #ปรับ bias
        deltaB = np.array(gradient_output).dot(-self.learningRate)
        self.bias[1] = self.bias[1] + deltaB
        deltaB = np.array(gradient_Layer2).dot(-self.learningRate)
        self.bias[0] = self.bias[0] + deltaB
        
        #ปรับ weight
        deltaW = np.array(deltaW)
        self.w_Layer = self.w_Layer + deltaW

    def train(self,in_,out_):
        self.forward(in_)
        self.backward(self.output,out_)
    def test(self,in_put):
        self.forward(in_put)
        return self.output 

def start(s,a,b,c,d,e,f,g,h=200):
    for j in range(h):
        for n in range(c+1,d+2):
            temp = []
            err = []
            for nn in range(1,14):
                colum = sheet.cell(row=n, column=nn)
                temp.append(colum.value)
            for nnn in range(15,19):
                colum = sheet.cell(row=n, column=nnn)
                err.append(colum.value)
            s.train(temp,err)
        if g == True:
            for n in range(e+1,f+2):
                temp = []
            err = []
            for nn in range(1,14):
                colum = sheet.cell(row=n, column=nn)
                temp.append(colum.value)
            for nnn in range(15,19):
                colum = sheet.cell(row=n, column=nnn)
                err.append(colum.value)
            s.train(temp,err)
    match = 0
    for m in range(a+1,b+2):
        temp = []
        for mm in range(1,14):
            colum = sheet.cell(row=m, column=mm)
            temp.append(colum.value)
        maxx = None
        for mmm in range(15,19):
            colum = sheet.cell(row=m, column=mmm)
            if colum.value == 1:
                maxx = mmm-15
        max_i = 0
        arr = []
        arr = s.test(temp)
        for an in range(len(arr)):
            if arr[an] > arr[max_i]:
                max_i = an
        if maxx == max_i:
            match += 1
    return match

def FindNode():
    average = []
    for l in range(110,210,10):
        temp = 0
        ss1 = NeuralNetwork_(-0.5,7)
        temp += start(ss1,1,75,76,750,None,None,False,l)
        ss2 = NeuralNetwork_(-0.5,7)
        temp += start(ss2,76,150,1,75,151,750,True,l)
        ss3 = NeuralNetwork_(-0.5,7)
        temp += start(ss3,151,225,1,150,226,750,True,l)
        ss4 = NeuralNetwork_(-0.5,7)
        temp += start(ss4,226,300,1,225,301,750,True,l)
        ss5 = NeuralNetwork_(-0.5,7)
        temp += start(ss5,301,375,1,300,376,750,True,l)
        ss6 = NeuralNetwork_(-0.5,7)
        temp += start(ss6,376,450,1,375,451,750,True,l)
        ss7 = NeuralNetwork_(-0.5,7)
        temp += start(ss7,451,525,1,450,526,750,True,l)
        ss8 = NeuralNetwork_(-0.5,7)
        temp += start(ss8,526,600,1,525,601,750,True,l)
        ss9 = NeuralNetwork_(-0.5,7)
        temp += start(ss9,601,675,1,600,676,750,True,l)
        ss10 = NeuralNetwork_(-0.5,7)
        temp += start(ss10,676,750,1,675,None,None,False,l)
        average.append(temp/10)
    return average

forPlot = [1,2,3,4,5,6,7,8,9,10]
tt = FindNode()
tt = np.array(tt)*4/3
plt.plot(forPlot, tt, label = "epoch")

plt.xlabel('epoch') 
plt.ylabel('percent')
plt.legend()
plt.show() 